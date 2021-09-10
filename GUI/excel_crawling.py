import requests
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook

wb = Workbook()
time = time.localtime()

headers = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Edg/91.0.864.64"}

def create_soup(url) :
    res = requests.get(url, headers=headers)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, "lxml")
    return soup

def create_paging_soup(paging_url) :
    res = requests.get(paging_url, headers=headers)
    res.raise_for_status()
    paging_soup = BeautifulSoup(res.text, "lxml")
    return paging_soup

def Crawling(ws, sid1_code, sid2_code, date) :
    Crawling(ws, sid1_code, sid2_code, date)

def theme_100_264(date) :   # 정치 > 청와대
    sid1_code = str(100)
    sid2_code = str(264)

    ws = wb.active
    ws.title = "정치_청와대"
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_100_265(date) :   # 정치 > 국회/정당
    sid1_code = str(100)
    sid2_code = str(265)

    ws = wb.create_sheet("정치_국회,정당")
    ws = wb["정치_국회,정당"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_100_268(date) :   # 정치 > 북한
    sid1_code = str(100)
    sid2_code = str(268)

    ws = wb.create_sheet("정치_북한")
    ws = wb["정치_북한"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_100_266(date) :   # 정치 > 행정
    sid1_code = str(100)
    sid2_code = str(266)

    ws = wb.create_sheet("정치_행정")
    ws = wb["정치_행정"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_100_267(date) :   # 정치 > 국방/외교
    sid1_code = str(100)
    sid2_code = str(267)

    ws = wb.create_sheet("정치_국방,외교")
    ws = wb["정치_국방,외교"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_100_269(date) :   # 정치 > 정치일반
    sid1_code = str(100)
    sid2_code = str(269)

    ws = wb.create_sheet("정치_정치일반")
    ws = wb["정치_정치일반"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_101_259(date) :   # 경제 > 금융
    sid1_code = str(101)
    sid2_code = str(259)

    ws = wb.create_sheet("경제_금융")
    ws = wb["경제_금융"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_101_258(date) :   # 경제 > 증권
    sid1_code = str(101)
    sid2_code = str(258)

    ws = wb.create_sheet("경제_증권")
    ws = wb["경제_증권"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_101_261(date) :   # 경제 > 산업/재계
    sid1_code = str(101)
    sid2_code = str(261)

    ws = wb.create_sheet("경제_산업,재계")
    ws = wb["경제_산업,재계"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_101_771(date) :   # 경제 > 중기/벤처
    sid1_code = str(101)
    sid2_code = str(771)

    ws = wb.create_sheet("경제_중기,벤처")
    ws = wb["경제_중기,벤처"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_101_260(date) :   # 경제 > 부동산
    sid1_code = str(101)
    sid2_code = str(260)

    ws = wb.create_sheet("경제_부동산")
    ws = wb["경제_부동산"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_101_262(date) :   # 경제 > 글로벌경제
    sid1_code = str(101)
    sid2_code = str(262)

    ws = wb.create_sheet("경제_글로벌경제")
    ws = wb["경제_글로벌경제"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_101_310(date) :   # 경제 > 생활경제
    sid1_code = str(101)
    sid2_code = str(310)

    ws = wb.create_sheet("경제_생활경제")
    ws = wb["경제_생활경제"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_101_263(date) :   # 경제 > 경제일반
    sid1_code = str(101)
    sid2_code = str(263)

    ws = wb.create_sheet("경제_경제일반")
    ws = wb["경제_경제일반"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_102_249(date) :   # 사회 > 사건사고
    sid1_code = str(102)
    sid2_code = str(249)

    ws = wb.create_sheet("사회_사건사고")
    ws = wb["사회_사건사고"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_102_250(date) :   # 사회 > 교육
    sid1_code = str(102)
    sid2_code = str(250)

    ws = wb.create_sheet("사회_교육")
    ws = wb["사회_교육"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_102_251(date) :   # 사회 > 노동
    sid1_code = str(102)
    sid2_code = str(251)

    ws = wb.create_sheet("사회_노동")
    ws = wb["사회_노동"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_102_254(date) :   # 사회 > 언론
    sid1_code = str(102)
    sid2_code = str(254)

    ws = wb.create_sheet("사회_언론")
    ws = wb["사회_언론"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_102_252(date) :   # 사회 > 환경
    sid1_code = str(102)
    sid2_code = str(252)

    ws = wb.create_sheet("사회_환경")
    ws = wb["사회_환경"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_102_59b(date) :   # 사회 > 인권/복지
    sid1_code = str(102)
    sid2_code = "59b"   # 문자열

    ws = wb.create_sheet("사회_인권,복지")
    ws = wb["사회_인권,복지"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_102_255(date) :   # 사회 > 식품/의료
    sid1_code = str(102)
    sid2_code = str(255)

    ws = wb.create_sheet("사회_식품,의료")
    ws = wb["사회_식품,의료"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_102_256(date) :   # 사회 > 지역
    sid1_code = str(102)
    sid2_code = str(256)

    ws = wb.create_sheet("사회_지역")
    ws = wb["사회_지역"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_102_276(date) :   # 사회 > 인물
    sid1_code = str(102)
    sid2_code = str(276)

    ws = wb.create_sheet("사회_인물")
    ws = wb["사회_인물"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_102_257(date) :   # 사회 > 사회일반
    sid1_code = str(102)
    sid2_code = str(257)

    ws = wb.create_sheet("사회_사회일반")
    ws = wb["사회_사회일반"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_103_241(date) :   # 생활/문화 > 건강정보
    sid1_code = str(103)
    sid2_code = str(241)

    ws = wb.create_sheet("생활,문화_건강정보")
    ws = wb["생활,문화_건강정보"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_103_239(date) :   # 생활/문화 > 자동차/시승기
    sid1_code = str(103)
    sid2_code = str(239)

    ws = wb.create_sheet("생활,문화_자동차,시승기")
    ws = wb["생활,문화_자동차,시승기"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_103_240(date) :   # 생활/문화 > 도로/교통
    sid1_code = str(103)
    sid2_code = str(240)

    ws = wb.create_sheet("생활,문화_도로,교통")
    ws = wb["생활,문화_도로,교통"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_103_237(date) :   # 생활/문화 > 여행/레저
    sid1_code = str(103)
    sid2_code = str(237)

    ws = wb.create_sheet("생활,문화_여행,레저")
    ws = wb["생활,문화_여행,레저"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_103_238(date) :   # 생활/문화 > 음식/맛집
    sid1_code = str(103)
    sid2_code = str(238)

    ws = wb.create_sheet("생활,문화_음식,맛집")
    ws = wb["생활,문화_음식,맛집"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_103_376(date) :   # 생활/문화 > 패션/뷰티
    sid1_code = str(103)
    sid2_code = str(376)

    ws = wb.create_sheet("생활,문화_패션,뷰티")
    ws = wb["생활,문화_패션,뷰티"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_103_242(date) :   # 생활/문화 > 공연/전시
    sid1_code = str(103)
    sid2_code = str(242)

    ws = wb.create_sheet("생활,문화_공연,전시")
    ws = wb["생활,문화_공연,전시"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_103_243(date) :   # 생활/문화 > 책
    sid1_code = str(103)
    sid2_code = str(243)

    ws = wb.create_sheet("생활,문화_책")
    ws = wb["생활,문화_책"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_103_244(date) :   # 생활/문화 > 종교
    sid1_code = str(103)
    sid2_code = str(244)

    ws = wb.create_sheet("생활,문화_종교")
    ws = wb["생활,문화_종교"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_103_248(date) :   # 생활/문화 > 날씨
    sid1_code = str(103)
    sid2_code = str(248)

    ws = wb.create_sheet("생활,문화_날씨")
    ws = wb["생활,문화_날씨"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_103_245(date) :   # 생활/문화 > 생활문화일반
    sid1_code = str(103)
    sid2_code = str(245)

    ws = wb.create_sheet("생활,문화_생활문화일반")
    ws = wb["생활,문화_생활문화일반"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_104_231(date) :   # 세계 > 아시아/호주
    sid1_code = str(104)
    sid2_code = str(231)

    ws = wb.create_sheet("세계_아시아,호주")
    ws = wb["세계_아시아,호주"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_104_232(date) :   # 세계 > 미국/중남미
    sid1_code = str(104)
    sid2_code = str(232)

    ws = wb.create_sheet("세계_미국,중남미")
    ws = wb["세계_미국,중남미"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_104_233(date) :   # 세계 > 유럽
    sid1_code = str(104)
    sid2_code = str(233)

    ws = wb.create_sheet("세계_유럽")
    ws = wb["세계_유럽"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_104_234(date) :   # 세계 > 중동/아프리카
    sid1_code = str(104)
    sid2_code = str(234)

    ws = wb.create_sheet("세계_중동,아프리카")
    ws = wb["세계_중동,아프리카"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_104_322(date) :   # 세계 > 세계일반
    sid1_code = str(104)
    sid2_code = str(322)

    ws = wb.create_sheet("세계_세계일반")
    ws = wb["세계_세계일반"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_105_731(date) :   # IT/과학 > 모바일
    sid1_code = str(105)
    sid2_code = str(731)

    ws = wb.create_sheet("IT,과학_모바일")
    ws = wb["IT,과학_모바일"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_105_226(date) :   # IT/과학 > 인터넷/SNS
    sid1_code = str(105)
    sid2_code = str(226)

    ws = wb.create_sheet("IT,과학_인터넷,SNS")
    ws = wb["IT,과학_인터넷,SNS"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_105_227(date) :   # IT/과학 > 통신/뉴미디어
    sid1_code = str(105)
    sid2_code = str(227)

    ws = wb.create_sheet("IT,과학_통신,뉴미디어")
    ws = wb["IT,과학_통신,뉴미디어"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_105_230(date) :   # IT/과학 > IT일반
    sid1_code = str(105)
    sid2_code = str(230)

    ws = wb.create_sheet("IT,과학_IT일반")
    ws = wb["IT,과학_IT일반"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_105_732(date) :   # IT/과학 > 보안/해킹
    sid1_code = str(105)
    sid2_code = str(732)

    ws = wb.create_sheet("IT,과학_보안,해킹")
    ws = wb["IT,과학_보안,해킹"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_105_283(date) :   # IT/과학 > 컴퓨터
    sid1_code = str(105)
    sid2_code = str(283)

    ws = wb.create_sheet("IT,과학_컴퓨터")
    ws = wb["IT,과학_컴퓨터"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_105_229(date) :   # IT/과학 > 게임/리뷰
    sid1_code = str(105)
    sid2_code = str(229)

    ws = wb.create_sheet("IT,과학_게임,리뷰")
    ws = wb["IT,과학_게임,리뷰"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def theme_105_228(date) :   # IT/과학 > 과학일반
    sid1_code = str(105)
    sid2_code = str(228)

    ws = wb.create_sheet("IT,과학_과학일반")
    ws = wb["IT,과학_과학일반"]
    ws.append(["index", "제목", "링크", "언론사"])

    Crawling(ws, sid1_code, sid2_code, date)

def scrape_news_excel(date) :
    theme_100_264(date) # 정치_청와대
    theme_100_265(date) # 정치_국회/정당
    theme_100_268(date) # 정치_북한
    theme_100_266(date) # 정치_행정
    theme_100_267(date) # 정치_국방/외교
    theme_100_269(date) # 정치_정치일반

    theme_101_259(date) # 경제_금융
    theme_101_258(date) # 경제_증권
    theme_101_261(date) # 경제_산업/재계
    theme_101_771(date) # 경제_중기/벤처
    theme_101_260(date) # 경제_부동산
    theme_101_262(date) # 경제_글로벌경제
    theme_101_310(date) # 경제_생활경제
    theme_101_263(date) # 경제_경제일반

    theme_102_249(date) # 사회_사건사고
    theme_102_250(date) # 사회_교육
    theme_102_251(date) # 사회_노동
    theme_102_254(date) # 사회_언론
    theme_102_252(date) # 사회_환경
    theme_102_59b(date) # 사회_인권/복지
    theme_102_255(date) # 사회_식품/의료
    theme_102_256(date) # 사회_지역
    theme_102_276(date) # 사회_인물
    theme_102_257(date) # 사회_사회일반

    theme_103_241(date) # 생활/문화_건강정보
    theme_103_239(date) # 생활/문화_자동차/시승기
    theme_103_240(date) # 생활/문화_도로/교통
    theme_103_237(date) # 생활/문화_여행/레저
    theme_103_238(date) # 생활/문화_음식/맛집
    theme_103_376(date) # 생활/문화_패션/뷰티
    theme_103_242(date) # 생활/문화_공연/전시
    theme_103_243(date) # 생활/문화_책
    theme_103_244(date) # 생활/문화_종교
    theme_103_248(date) # 생활/문화_날씨
    theme_103_245(date) # 생활/문화_생활문화일반

    theme_104_231(date) # 세계_아시아/호주
    theme_104_232(date) # 세계_미국/중남미
    theme_104_233(date) # 세계_유럽
    theme_104_234(date) # 세계_중동/아프리카
    theme_104_322(date) # 세계_세계일반

    theme_105_731(date) # IT/과학_모바일
    theme_105_226(date) # IT/과학_인터넷/SNS
    theme_105_227(date) # IT/과학_통신/뉴미디어
    theme_105_230(date) # IT/과학_IT일반
    theme_105_732(date) # IT/과학_보안/해킹
    theme_105_283(date) # IT/과학_컴퓨터
    theme_105_229(date) # IT/과학_게임/리뷰
    theme_105_228(date) # IT/과학_과학일반

    wb.save("Crawiling을 진행한 시각 {}.{}.{} {}시{}분{}초.xlsx".format(time.tm_year, time.tm_mon, time.tm_mday, time.tm_hour, time.tm_min, time.tm_sec))

if __name__ == "__main__" :
    scrape_news_excel() # 뉴스 정보 가져오기
